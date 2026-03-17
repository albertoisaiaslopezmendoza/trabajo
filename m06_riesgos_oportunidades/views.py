import base64
from django.shortcuts import render, redirect, get_object_or_404
from django.core.files.base import ContentFile
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

from .models import RiesgoOportunidad, FirmaRiesgo
from .forms import RiesgoOportunidadForm


def dashboard_m06(request):
    registros = RiesgoOportunidad.objects.all().order_by('-fecha_registro')

    if request.method == 'POST':
        form = RiesgoOportunidadForm(request.POST)
        if form.is_valid():
            riesgo = form.save()

            # Procesar la firma web si se dibujó algo
            firma_b64 = request.POST.get('firma_base64')
            firmante_nombre = request.POST.get('revisado_por') or riesgo.responsable or 'N/D'

            if firma_b64:
                # El formato base64 de JS viene como: "data:image/png;base64,iVBORw0KGgo..."
                format, imgstr = firma_b64.split(';base64,')
                ext = format.split('/')[-1]
                nombre_archivo = f"firma_m06_riesgo_{riesgo.id}.{ext}"
                data = ContentFile(base64.b64decode(imgstr), name=nombre_archivo)

                # Crear el registro de la firma y asociarlo al riesgo
                FirmaRiesgo.objects.create(
                    riesgo=riesgo,
                    nombre=firmante_nombre,
                    rol='Revisión/Aprobación',
                    firma_imagen=data
                )
            return redirect('m06_riesgos_oportunidades:m06_dashboard')
    else:
        form = RiesgoOportunidadForm(initial={'probabilidad': 1, 'impacto': 1})

    context = {
        'registros': registros,
        'form': form,
    }
    return render(request, 'm06_riesgos_oportunidades/dashboard.html', context)


def exportar_excel_m06(request):
        # Obtener registros
        registros = RiesgoOportunidad.objects.all().order_by('-fecha_registro')

        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz de Riesgos"

        # Encabezados
        headers = [
            "ID", "Proceso", "Tipo", "Estado", "Prob(1-5)", "Impacto(1-5)",
            "Nivel(PxI)", "Clasificación", "Descripción", "Causa", "Consecuencia",
            "Controles actuales", "Acción planificada", "Responsable",
            "Fecha objetivo", "Evidencia", "Creado por", "Revisado por", "Fecha Registro"
        ]
        ws.append(headers)

        # Llenar datos
        for r in registros:
            # Dar formato a la fecha para que no de error en Excel si está vacía
            fecha_obj = r.fecha_objetivo.strftime('%Y-%m-%d') if r.fecha_objetivo else "N/D"
            fecha_reg = r.fecha_registro.strftime('%Y-%m-%d %H:%M') if r.fecha_registro else "N/D"

            # Validamos los campos que suelen quedar vacíos
            evidencia_val = r.evidencia if r.evidencia else "Sin evidencia registrada"
            creado_por_val = r.creado_por if r.creado_por else "No especificado"
            revisado_por_val = r.revisado_por if r.revisado_por else "Pendiente de revisión"

            ws.append([
                r.id, r.proceso, r.tipo, r.estado, r.probabilidad, r.impacto,
                r.nivel, r.clasificacion, r.descripcion, r.causa, r.consecuencia,
                r.control_actual, r.accion_plan, r.responsable or "N/D", fecha_obj,
                evidencia_val, creado_por_val, revisado_por_val, fecha_reg
            ])

        # Ajustar anchos de columna (opcional para que se vea bien)
        widths = [6, 18, 12, 12, 10, 11, 10, 13, 35, 25, 25, 25, 25, 16, 14, 25, 20, 20, 16]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Congelar la primera fila y agregar filtros
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # Preparar respuesta HTTP con el archivo Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_archivo = f"M06_MatrizRiesgos_{timestamp}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        wb.save(response)

        return response